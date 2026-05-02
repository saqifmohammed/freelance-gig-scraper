import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
from scrapers import Job


class ExcelStorage:
    def __init__(self):
        self.filename = "job_leads.xlsx"
        self.filepath = os.path.join(os.path.dirname(__file__), self.filename)
        self.header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.alternate_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")

    def save_jobs(self, jobs: list[Job]):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        sheet_name = f"Run_{timestamp}"

        if os.path.exists(self.filepath):
            wb = openpyxl.load_workbook(self.filepath)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        headers = ["Platform", "Job Title", "Description", "Budget", "URL", "Posted Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row, job in enumerate(jobs, 2):
            ws.cell(row=row, column=1, value=job.platform)
            ws.cell(row=row, column=2, value=job.title)
            ws.cell(row=row, column=3, value=job.description or "N/A")
            ws.cell(row=row, column=4, value=job.budget or "N/A")
            ws.cell(row=row, column=5, value=job.url)
            ws.cell(row=row, column=6, value=job.posted_date or "N/A")

            if row % 2 == 0:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = self.alternate_fill

        for col in range(1, 7):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

        ws.column_dimensions['C'].width = 60

        wb.save(self.filepath)
        return self.filepath

    def get_all_jobs(self) -> list:
        jobs = []
        if not os.path.exists(self.filepath):
            return jobs

        wb = openpyxl.load_workbook(self.filepath)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    jobs.append({
                        "sheet": sheet,
                        "platform": row[0],
                        "title": row[1],
                        "description": row[2],
                        "budget": row[3],
                        "url": row[4],
                        "posted_date": row[5]
                    })

        wb.close()
        return jobs